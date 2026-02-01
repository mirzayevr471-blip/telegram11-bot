import asyncio
import os
from aiogram import Bot, Dispatcher, types
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile
from openpyxl import Workbook, load_workbook

# ================== SOZLAMALAR ==================
TOKEN = ""
ADMIN_IDS = []
ADMIN_GROUPS = []
FILE_NAME = "users.xlsx"

bot = Bot(token=TOKEN)
dp = Dispatcher()

# ================== EXCEL INIT ==================
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Ism", "Familiya", "Telefon",
        "Viloyat", "Tuman", "Mahalla",
        "Yosh", "Telegram ID"
    ])
    wb.save(FILE_NAME)

# ================== FSM ==================
class LangState(StatesGroup):
    lang = State()

class Register(StatesGroup):
    ism = State()
    familya = State()
    telefon = State()
    viloyat = State()
    tuman = State()
    mahalla = State()
    yosh = State()

class DeleteUser(StatesGroup):
    telegram_id = State()

# ================== MATNLAR ==================
TEXTS = {
    "uz": {
        "ask_name": "👤 Ismingizni kiriting:",
        "done": "✅ Siz muvaffaqiyatli ro‘yxatdan o‘tdingiz!\n📞 Admin siz bilan bog‘lanadi."
    },
    "ru": {
        "ask_name": "👤 Введите имя:",
        "done": "✅ Вы успешно зарегистрировались!"
    },
    "en": {
        "ask_name": "👤 Enter your name:",
        "done": "✅ You are registered!"
    }
}

# ================== /start ==================
@dp.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    await state.clear()
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🇺🇿 O‘zbekcha")],
            [KeyboardButton(text="🇷🇺 Русский")],
            [KeyboardButton(text="🇬🇧 English")]
        ],
        resize_keyboard=True
    )
    await message.answer("🌐 Tilni tanlang:", reply_markup=kb)
    await state.set_state(LangState.lang)

@dp.message(LangState.lang)
async def set_language(message: types.Message, state: FSMContext):
    if "O‘zbek" in message.text:
        lang = "uz"
    elif "Рус" in message.text:
        lang = "ru"
    elif "English" in message.text:
        lang = "en"
    else:
        return

    await state.update_data(lang=lang)
    await message.answer(TEXTS[lang]["ask_name"], reply_markup=ReplyKeyboardRemove())
    await state.set_state(Register.ism)

# ================== REGISTER ==================
@dp.message(Register.ism)
async def get_ism(message: types.Message, state: FSMContext):
    await state.update_data(ism=message.text)
    await message.answer("👤 Familyangizni kiriting:")
    await state.set_state(Register.familya)

@dp.message(Register.familya)
async def get_familya(message: types.Message, state: FSMContext):
    await state.update_data(familya=message.text)
    kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📞 Telefon yuborish", request_contact=True)]],
        resize_keyboard=True
    )
    await message.answer("📞 Telefon yuboring:", reply_markup=kb)
    await state.set_state(Register.telefon)

@dp.message(Register.telefon)
async def get_phone(message: types.Message, state: FSMContext):
    if not message.contact:
        return
    await state.update_data(telefon=message.contact.phone_number)
    await message.answer("🏠 Viloyat:")
    await state.set_state(Register.viloyat)

@dp.message(Register.viloyat)
async def get_viloyat(message: types.Message, state: FSMContext):
    await state.update_data(viloyat=message.text)
    await message.answer("🏘 Tuman:")
    await state.set_state(Register.tuman)

@dp.message(Register.tuman)
async def get_tuman(message: types.Message, state: FSMContext):
    await state.update_data(tuman=message.text)
    await message.answer("🏡 Mahalla:")
    await state.set_state(Register.mahalla)

@dp.message(Register.mahalla)
async def get_mahalla(message: types.Message, state: FSMContext):
    await state.update_data(mahalla=message.text)
    await message.answer("🎂 Yoshingiz (10–20):")
    await state.set_state(Register.yosh)

@dp.message(Register.yosh)
async def get_yosh(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        return
    yosh = int(message.text)
    if not 10 <= yosh <= 20:
        return

    await state.update_data(yosh=yosh)
    data = await state.get_data()

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([
        data["ism"], data["familya"], data["telefon"],
        data["viloyat"], data["tuman"], data["mahalla"],
        data["yosh"], message.from_user.id
    ])
    wb.save(FILE_NAME)

    await bot.send_message(ADMIN_GROUPS[0], f"🆕 Yangi foydalanuvchi\n🆔 {message.from_user.id}")
    await message.answer(TEXTS[data["lang"]]["done"])
    await state.clear()

# ================== ADMIN PANEL ==================
@dp.message(Command("admin"))
async def admin_panel(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Foydalanuvchilar soni")],
            [KeyboardButton(text="📥 Excel yuklab olish")],
            [KeyboardButton(text="🗑 Foydalanuvchini o‘chirish")]
        ],
        resize_keyboard=True
    )
    await message.answer("🧑‍⚖️ Admin panel", reply_markup=kb)

@dp.message(lambda m: m.text == "📊 Foydalanuvchilar soni")
async def count_users(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    await message.answer(f"👥 Jami: {ws.max_row - 1}")

@dp.message(lambda m: m.text == "📥 Excel yuklab olish")
async def send_excel(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    await message.answer_document(FSInputFile(FILE_NAME))

@dp.message(lambda m: m.text == "🗑 Foydalanuvchini o‘chirish")
async def delete_start(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await message.answer("🆔 Telegram ID yuboring:")
    await state.set_state(DeleteUser.telegram_id)

@dp.message(DeleteUser.telegram_id)
async def delete_finish(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        return

    target = int(message.text)
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=8).value == target:
            ws.delete_rows(row)
            wb.save(FILE_NAME)
            await message.answer("✅ O‘chirildi")
            await state.clear()
            return

    await message.answer("❌ Topilmadi")
    await state.clear()

# ================== RUN ==================
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())

